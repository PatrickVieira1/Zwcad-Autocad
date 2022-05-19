from pyzwcad import ZwCAD, APoint
import win32com.client
import os
import time
from openpyxl import Workbook
import regex as re
import pyautogui

acad = win32com.client.Dispatch("ZWCAD.Application")
root = r'''//TX-FS/usuarios/Engenharia/Pedidos'''
wb = Workbook()
ws = wb.active
ws.title = 'Zwcad'
#Lista de subpastas a serem procuradas por desenhos
l = 24
lista = []
inicio = 'BR13'
for k in range(14):
    lista.append(inicio+str(l))
    l+=1  # incrementa o contador
#228 - 285
ListaExcluidos = []
#Lista de subpastas a serem ignoradas
for i in range(228, 285):
    ListaExcluidos.append(str(i))
for i in range(800,824):
    ListaExcluidos.append(str(i))
for i in range(0,10):
    ListaExcluidos.append(inicio+str('0'+str(i)))
for i in range(10,24):
    ListaExcluidos.append(inicio+str(i))

i = 0 
j = 0

for dirpath, dirnames, filenames in os.walk(root):
    #iteração para ignorar cada subpasta conforme lista de excluidos
    dirnames[:] = [d for d in dirnames if d not in ListaExcluidos]
    #try-except para evitar erros de arquivos não encontrados
    try:
        if (dirpath.replace(root, '')).split('\\')[1] in lista:
            for filename in filenames:
                if filename.endswith('.dwg') and re.search(r'-0000*', filename)!= None:
                    j += 1
                    i += 1
                    print(filename)
                    print('aqui esta o caminho' + os.path.join(dirpath, filename))
                    acad.Documents.Open(os.path.join(dirpath,filename))
                    time.sleep(2)
                    for entity in acad.ActiveDocument.PaperSpace:
                        name = entity.EntityName
                        if name == 'AcDbBlockReference':
                            HasAttributes = entity.HasAttributes
                            if HasAttributes:
                                for attrib in entity.GetAttributes():
                                    ws.cell(row=j, column=1).value = filename
                                    ws.cell(column=2, row=j, value=os.path.join(dirpath, filename))
                                    if attrib.TagString == 'DESCRIÇÃO':
                                        #print(attrib.TextString)

                                        ws.cell(row=j, column=i+10).value = attrib.TextString
                                        
                                        i += 1
                                    elif attrib.TagString == 'CP':
                                        ws.cell(row=j, column=3).value = attrib.TextString
                                    elif attrib.TagString == 'MODELO':
                                        ws.cell(row=j, column=4).value = attrib.TextString
                                    elif attrib.TagString == 'EQUIP':
                                        ws.cell(row=j, column=5).value = attrib.TextString
                                    elif attrib.TagString == 'NDESE':
                                        ws.cell(row=j, column=6).value = attrib.TextString
                                    elif attrib.TagString == 'CLIENTE':
                                        ws.cell(row=j, column=7).value = attrib.TextString
                                    elif attrib.TagString == 'R':
                                        ws.cell(row=j, column=9).value = attrib.TextString
                                    elif attrib.TagString == 'MATERIAL' and re.search(r'carc',attrib.TextString, re.IGNORECASE)!= None:
                                        ws.cell(row=j, column=10).value = attrib.TextString
                                        print(attrib.TextString)
 
                                    #print("  {}: {}".format(attrib.TagString, attrib.TextString))

                        elif name == 'AcDbMText':
                            #re.search(pattern, string, flags=0)
                            #re.escape para fugir de caracteres como /n ou algo do tipo
                            #re.search vs re.match - o primeiro retorna o primeiro resultado, o segundo retorna se o texto começa com o padrão
                            regex = re.search('notas',re.escape(entity.TextString),  re.IGNORECASE)
                            if regex:
                                Notas = entity.TextString.replace('\P', ' - ')
                                ws.cell(row=j, column=8).value = str(Notas)
                pyautogui.moveTo(300,300)
                acad.Documents.Close()
                #time.sleep(2)
                #Retorno pro inicio da coluna nas iterações do excel
                i = 0
    except:
        pass

wb.save(filename='Zwcad.xlsx')

